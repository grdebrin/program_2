import openpyxl

def process_steps(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    columns_to_process = ["Шаги", "Ожидаемый результат"]

    for column_name in columns_to_process:
        found_column = None
        for column in sheet.iter_cols():
            if column[0].value == column_name:
                found_column = column
                break

        if found_column:
            i = 0
            index = 1
            while i < len(found_column):
                cell = found_column[i]
                if cell.value is not None:
                    main_step = cell
                    sub_steps = [f"{index}. \"{cell.value}\""]
                    index += 1
                    i += 1
                    if i < len(found_column):
                        cell = found_column[i]
                        while cell.value is not None:
                            sub_steps.append(f"{index}. \"{cell.value}\"")
                            cell.value = None
                            index += 1
                            i += 1
                            if i < len(found_column):
                                cell = found_column[i]
                            else:
                                break
                        main_step.value = '\n'.join(sub_steps)
                        if main_step.row > 1:  # Check if the cell is not in the first row
                            main_step_coord = main_step.coordinate
                            destination = sheet.cell(row=main_step.row - 1, column=main_step.column)
                            destination.value = main_step.value
                            main_step.value = None  # Clear the main step cell after moving the value
                else:
                    index = 1
                    i += 1

    workbook.save(file_path)

if __name__ == "__main__":
    file_path = "your_path"
    process_steps(file_path)
